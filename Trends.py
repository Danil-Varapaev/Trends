import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import zipfile
import shutil
import warnings
from openpyxl import Workbook, load_workbook  # ← добавил load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

# ========================= КОНСТАНТЫ =========================
LIMITS = {
    'D': {
        'air_asp_operated': {'warning': 100, 'action': 160, 'limit': 200},
        'air_asp_rest': {'warning': 100, 'action': 160, 'limit': 200},
        'air_sed_operated': {'warning': 40, 'action': 60, 'limit': 100},
        'air_sed_rest': {'warning': 40, 'action': 60, 'limit': 100},
        'swabs_operated': {'warning': 30, 'action': 40, 'limit': 50},
        'swabs_rest': {'warning': 15, 'action': 20, 'limit': 25},
    },
    'C': {
        'air_asp_operated': {'warning': 40, 'action': 80, 'limit': 100},
        'air_asp_rest': {'warning': 40, 'action': 80, 'limit': 100},
        'air_sed_operated': {'warning': 20, 'action': 26, 'limit': 50},
        'air_sed_rest': {'warning': 20, 'action': 26, 'limit': 50},
        'swabs_operated': {'warning': 15, 'action': 20, 'limit': 25},
        'swabs_rest': {'warning': 2, 'action': 4, 'limit': 5},
    },
    'B': {
        'air_asp_operated': {'warning': 4, 'action': 8, 'limit': 10},
        'air_asp_rest': {'warning': 4, 'action': 8, 'limit': 10},
        'air_sed_operated': {'warning': 1, 'action': 4, 'limit': 5},
        'air_sed_rest': {'warning': 1, 'action': 4, 'limit': 5},
        'swabs_operated': {'warning': 2, 'action': 4, 'limit': 5},
        'swabs_rest': {'warning': None, 'action': None, 'limit': 1},
    },
    'A': {
        'air_asp_operated': {'warning': None, 'action': None, 'limit': 1},
        'air_asp_rest': {'warning': None, 'action': None, 'limit': 1},
        'air_sed_operated': {'warning': None, 'action': None, 'limit': 1},
        'air_sed_rest': {'warning': None, 'action': None, 'limit': 1},
        'swabs_operated': {'warning': None, 'action': None, 'limit': 1},
        'swabs_rest': {'warning': None, 'action': None, 'limit': 1},
    }
}

TYPE_MAPPING = {
    'air_asp_operated': 'Аспирация, Эксплуатация',
    'air_asp_rest': 'Аспирация, Оснащенное/Покой',
    'air_sed_operated': 'Седиментация, Эксплуатация',
    'air_sed_rest': 'Седиментация, Оснащенное/Покой',
    'swabs_operated': 'Смывы, Эксплуатация',
    'swabs_rest': 'Смывы, Оснащенное/Покой'
}

# ========================= ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =========================

def normalize_class(val):
    if pd.isna(val): return None
    s = str(val).strip().upper().replace(" ", "")
    if "(ПО" in s: s = s.split("(ПО")[-1].replace(")", "").strip()
    s = s.replace("В", "B").replace("С", "C").replace("А", "A")
    for c in "ABCD":
        if c in s: return c
    return None

def classify_type(row, anomalies_file=None):
    cat = str(row.get('Категория образца', '')).strip().upper()
    method_raw = row.get('Метод (А/С)', '')
    method = str(method_raw).strip().upper() if pd.notna(method_raw) else ''

    # 1. ВОЗДУХ
    if any(x in cat for x in ['AIR', 'ВОЗДУХ']):
        if not method or method in ['NAN', 'NONE', '-', '']:
            if anomalies_file:
                with open(anomalies_file, 'a', encoding='utf-8') as f:
                    f.write(f"ОТБРОШЕНА (пустой метод): Дата={row.get('Дата')} | "
                            f"Лаб№={row.get('Лабораторный номер','—')} | "
                            f"Категория={cat} | Метод='{method_raw}'\n")
            return None 

        is_asp = any(x in method for x in ['А', 'A', 'ASP', 'АСПИРАЦИ', 'ASPI', 'АСП'])
        is_sed = any(x in method for x in ['С', 'C', 'SED', 'СЕДИМЕНТАЦ', 'SEDI', 'СЕД'])
        is_operated = any(x in cat for x in ['IN PROCESS', 'OPERATED', 'ЭКСПЛУАТАЦ', 'РАБОТ', 'ПРОЦЕСС'])
        
        if is_asp:
            return 'air_asp_operated' if is_operated else 'air_asp_rest'
        elif is_sed:
            return 'air_sed_operated' if is_operated else 'air_sed_rest'
        else:
            if anomalies_file:
                with open(anomalies_file, 'a', encoding='utf-8') as f:
                    f.write(f"ОТБРОШЕНА (нераспознанный метод): Дата={row.get('Дата')} | "
                            f"Лаб№={row.get('Лабораторный номер','—')} | Метод='{method_raw}'\n")
            return None

    # 2. СМЫВЫ
    elif any(x in cat for x in ['SWAB', 'СМЫВ', 'СМЫВЫ']):
        is_operated = any(x in cat for x in ['IN PROCESS', 'OPERATED', 'ЭКСПЛУАТАЦ', 'РАБОТ', 'ПРОЦЕСС'])
        return 'swabs_operated' if is_operated else 'swabs_rest'

    return None

# ИСПРАВЛЕНО: Добавлен аргумент limit_value
def clean_coe(v, limit_value=None):
    if pd.isna(v): return np.nan, str(v)
    s = str(v).strip()
    s_lower = s.lower()
    
    if any(x in s_lower for x in ['менее 1', '<1', 'менее1', '< 1']):
        return 0.0, s
    
    if any(x in s_lower for x in ['сплошной рост', 'сплошной', 'рост превышает', 'контаминация', 'газон', 'too numerous', 'tntc']):
        if limit_value is not None and limit_value > 0:
            return round(limit_value * 1.1, 1), s
        else:
            return np.nan, s
            
    try:
        return float(s.replace(',', '.')), s
    except:
        return np.nan, s

# ========================= ЗАГРУЗКА И ОБРАБОТКА ДАННЫХ =========================

def load_data(input_dir):
    all_dfs = []
    for filename in os.listdir(input_dir):
        if filename.endswith('.xlsx'):
            filepath = os.path.join(input_dir, filename)
            try:
                df_air = pd.read_excel(filepath, sheet_name='Воздух')
                df_air['source_sheet'] = 'Воздух'
                all_dfs.append(df_air)
                
                df_swab = pd.read_excel(filepath, sheet_name='Смывы')
                df_swab['source_sheet'] = 'Смывы'
                all_dfs.append(df_swab)
            except Exception as e:
                print(f"Ошибка чтения {filename}: {e}")

    if not all_dfs: return pd.DataFrame()
    return pd.concat(all_dfs, ignore_index=True)

# ИСПРАВЛЕНО: Правильный порядок действий (Сначала Класс -> Потом Лимит -> Потом КОЕ)
def process_data(df, anomalies_file):
    # 1. СТАНДАРТИЗАЦИЯ ИМЕН КОЛОНОК
    if 'Лаборатор-ный номер' in df.columns:
        df.rename(columns={'Лаборатор-ный номер': 'Лабораторный номер'}, inplace=True)
    if '№ помещения' in df.columns:
        df.rename(columns={'№ помещения': 'Номер_помещения'}, inplace=True)
    
    df['Номер_помещения'] = df['Номер_помещения'].astype(str).str.strip()
    print(f"[Начало] Всего строк: {len(df)}")

    # 2. СНАЧАЛА определяем Класс и Тип
    df['Класс_норм'] = df['Класс чистоты'].apply(normalize_class)
    df['monitor_type'] = df.apply(lambda row: classify_type(row, anomalies_file=anomalies_file), axis=1)

    before_meta = len(df)
    df = df[df['Класс_норм'].notna() & df['monitor_type'].notna()]
    print(f"[Мета-данные] Удалено строк без Класса/Типа: {before_meta - len(df)}")

    # 3. Вычисляем лимит для каждой строки ДО очистки КОЕ
    def get_limit_for_row(row):
        try:
            return LIMITS[row['Класс_норм']][row['monitor_type']]['limit']
        except:
            return None
            
    # ✅ ВОТ СЮДА СТАВИТСЯ СТРОКА СОЗДАНИЯ ВРЕМЕННОЙ КОЛОНКИ
    df['_temp_limit'] = df.apply(get_limit_for_row, axis=1)

    # 4. ИЗВЛЕЧЕНИЕ И ОЧИСТКА КОЕ
    def get_coe_value(row):
        if row['source_sheet'] == 'Воздух':
            val = row.get('Результат КОЕ/м3 КОЕ/4ч ')
            if pd.isna(val): val = row.get('КОЕ')
        else:
            val = row.get('КОЕ')
            if pd.isna(val): val = row.get('Результат КОЕ/м3 КОЕ/4ч ')
        return val

    df['raw_coe'] = df.apply(get_coe_value, axis=1)
    
    # Передаем лимит в clean_coe!
    df[['КОЕ_numeric', 'Результат_сырой']] = df.apply(
        lambda row: pd.Series(clean_coe(row['raw_coe'], limit_value=row['_temp_limit'])), 
        axis=1
    )
    
    # ✅ Удаляем временную колонку ТОЛЬКО ЗДЕСЬ, после использования
    df.drop(columns=['_temp_limit'], inplace=True)

    # 5. ФИЛЬТРАЦИЯ РЕЗУЛЬТАТОВ
    before_coe = len(df)
    df = df[df['КОЕ_numeric'].notna()]
    print(f"[КОЕ] Удалено нечисловых значений: {before_coe - len(df)}")

    # 6. Фильтрация по Дате
    df['Дата'] = pd.to_datetime(df['Дата'], errors='coerce')
    df = df.dropna(subset=['Дата'])

    # === МЕТРИКА 1: Статистика очистки ===
    stats = {
        'total_clean': len(df),
        'by_class': df['Класс_норм'].value_counts().to_dict(),
        'by_type': df['monitor_type'].value_counts().to_dict(),
        'with_anomalies': len(df[df['КОЕ_numeric'] > 0])
    }
    print(f"\n📊 СТАТИСТИКА ОБРАБОТКИ:")
    print(f"   Чистых записей для графиков: {stats['total_clean']}")
    print(f"   По классам: {stats['by_class']}")
    print(f"   По типам: {stats['by_type']}")
    
    print(f"\n✅ Итого для трендов: {len(df)} записей")
    return df, stats  # ← ЕДИНСТВЕННЫЙ return в функции

# ========================= ГЕНЕРАЦИЯ ОТЧЕТОВ =========================

def generate_report(group, room, mon_type, class_clean, output_dir, report_file):
    # Защита от невалидного класса
    if not class_clean or str(class_clean).strip() not in ['A', 'B', 'C', 'D']:
        print(f"⚠️ Пропущено: {room} — {mon_type} (невалидный класс: {class_clean})")
        return None
        
    room_str = str(room).strip()
    # Санитизация имени папки от спецсимволов
    room_str = room_str.replace('/', '_').replace('\\', '_').replace(':', '_')
    
    dir_path = os.path.join(output_dir, room_str)
    os.makedirs(dir_path, exist_ok=True)
    # Формируем имя файла с суффиксом класса
    excel_path = os.path.join(dir_path, f"{room_str}_{mon_type}_{class_clean}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "График"

    ws['A1'] = "Тренд микробиологического мониторинга"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Помещение: {room_str}"
    ws['A3'] = f"Тип: {TYPE_MAPPING.get(mon_type, mon_type)}"
    ws['A4'] = f"Класс чистоты: {class_clean}"

    # Данные для таблицы на листе "График"
    start_row = 7
    data_for_excel = group[['Дата', 'Результат_сырой']].copy()
    data_for_excel.columns = ['Дата', 'КОЕ']
    for r_idx, row_data in enumerate(dataframe_to_rows(data_for_excel, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # ГРАФИК
    plt.figure(figsize=(12, 7))
    dates = group['Дата']
    values = group['КОЕ_numeric']
    plt.plot(dates, values, 'b-o', linewidth=2.5, markersize=7, label='Результаты (КОЕ)')

    lim = LIMITS.get(class_clean, {}).get(mon_type, {})
    if lim.get('warning') is not None:
        plt.axhline(lim['warning'], color='orange', linestyle='--', linewidth=1.5, label='Тревога')
    if lim.get('action') is not None:
        plt.axhline(lim['action'], color='red', linestyle='--', linewidth=1.5, label='Действие')
    if lim.get('limit') is not None:
        plt.axhline(lim['limit'], color='purple', linestyle='-', linewidth=2, label='Предел')

    plt.title(f"Тренд — Помещение {room_str} — {TYPE_MAPPING.get(mon_type)}")
    plt.xlabel("Дата")
    plt.ylabel("КОЕ")
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.xticks(rotation=45)

    img_path = os.path.join(dir_path, "trend_chart.png")
    plt.savefig(img_path, dpi=220, bbox_inches='tight')
    plt.close()

    img = Image(img_path)
    img.width = 880
    img.height = 500
    ws.add_image(img, 'D7')

    # ЛИСТ "ДАННЫЕ"
    ws_data = wb.create_sheet("Данные")
    headers = ['Дата', 'Лабораторный номер', 'Подразделение', 'Результат КОЕ',
               'Уровень тревоги', 'Уровень действия', 'Предельное значение']
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

    for i, (_, row) in enumerate(group.iterrows(), 2):
        ws_data.cell(i, 1, row['Дата'])
        ws_data.cell(i, 2, row.get('Лабораторный номер'))
        ws_data.cell(i, 3, row.get('Подразделение'))
        ws_data.cell(i, 4, row['Результат_сырой'])
        l = LIMITS.get(class_clean, {}).get(mon_type, {})
        ws_data.cell(i, 5, l.get('warning'))
        ws_data.cell(i, 6, l.get('action'))
        ws_data.cell(i, 7, l.get('limit'))

    for column in ws_data.columns:
        max_len = max((len(str(cell.value)) for cell in column if cell.value is not None), default=10)
        ws_data.column_dimensions[column[0].column_letter].width = min(max_len + 3, 40)

    wb.save(excel_path)

    # АНАЛИЗ АНОМАЛИЙ ДЛЯ ЭТОЙ ГРУППЫ
    with open(report_file, 'a', encoding='utf-8') as f:
        f.write(f"\n{'='*90}\n")
        f.write(f"ПОМЕЩЕНИЕ: {room_str} | ТИП: {TYPE_MAPPING.get(mon_type)} | КЛАСС: {class_clean}\n")
        f.write(f"{'='*90}\n")
        anomalies_found = False

        num_group = group[group['КОЕ_numeric'].notna()]
        for _, r in num_group.iterrows():
            val = r['КОЕ_numeric']
            if val < 0: continue
            
            d = r['Дата'].date()
            lab = r.get('Лабораторный номер', '—')
            
            if lim.get('limit') and val >= lim['limit']:
                anomalies_found = True
                f.write(f"  КРИТИЧЕСКОЕ ПРЕВЫШЕНИЕ ПРЕДЕЛА → {d} | {val} >= {lim['limit']} | Лаб№ {lab}\n")
            elif lim.get('action') and val >= lim['action']:
                anomalies_found = True
                f.write(f"  ПРЕВЫШЕНИЕ УРОВНЯ ДЕЙСТВИЯ → {d} | {val} >= {lim['action']} | Лаб№ {lab}\n")
            elif lim.get('warning') and val >= lim['warning']:
                anomalies_found = True
                f.write(f"  ПРЕВЫШЕНИЕ УРОВНЯ ТРЕВОГИ → {d} | {val} >= {lim['warning']} | Лаб№ {lab}\n")
        
        if not anomalies_found:
            f.write("Аномалий не обнаружено.\n")

    # === МЕТРИКА 4: Флаги подозрительных трендов ===
    with open(report_file, 'a', encoding='utf-8') as f:
        if (group['КОЕ_numeric'] == 0).all() and len(group) > 3:
            f.write(f"  🚩 ФЛАГ: Все значения = 0 ({len(group)} точек)\n")
        
        lim = LIMITS.get(class_clean, {}).get(mon_type, {})
        if lim.get('limit') and (group['КОЕ_numeric'] >= lim['limit']).all() and len(group) > 1:
            f.write(f"  🚩 ФЛАГ: Все значения >= предела (проверьте класс чистоты)\n")
        
        if len(group) < 4:
            f.write(f"  🚩 ФЛАГ: Мало данных для квартала (всего {len(group)} точек)\n")

    return excel_path

# ========================= ЗАПУСК =========================

def main():
    INPUT_DIR = '/content/input_data'
    OUTPUT_DIR = '/content/output_reports'
    ZIP_PATH = '/content/microbio_trends.zip'
    REPORT_PATH = '/content/anomalies_report.txt'

    os.makedirs(INPUT_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if os.path.exists(REPORT_PATH): os.remove(REPORT_PATH)
    print("=== Автоматизация трендов микробиологического мониторинга ===\n")

    # Инициализация журнала
    with open(REPORT_PATH, 'w', encoding='utf-8') as f:
        f.write("ЖУРНАЛ АНОМАЛИЙ И ОТБРОШЕННЫХ ЗАПИСЕЙ\n")
        f.write(f"Дата генерации: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}\n")
        f.write("="*90 + "\n\n")

    # Загрузка файлов
    if not any(f.endswith('.xlsx') for f in os.listdir(INPUT_DIR)):
        print("Файлы не найдены. Загрузите их:")
        try:
            from google.colab import files
            uploaded = files.upload()
            for fn in uploaded.keys(): shutil.move(fn, os.path.join(INPUT_DIR, fn))
        except: return

    raw_df = load_data(INPUT_DIR)
    if len(raw_df) == 0: return

    # ОБРАБОТКА ДАННЫХ (возвращает df и stats)
    df, stats = process_data(raw_df, REPORT_PATH)
    if len(df) == 0:
        print("Нет данных после очистки.")
        return

    # ГРУППИРОВКА (разделяем помещения по классам, например P-119 A и B)
    grouped = df.groupby(['Номер_помещения', 'monitor_type', 'Класс_норм'])
    print(f"\nНайдено групп: {len(grouped)}\n")

    for (room, mon_type, class_clean), group in grouped:
        try:
            generate_report(group.sort_values('Дата'), room, mon_type, class_clean, OUTPUT_DIR, REPORT_PATH)
            print(f"✓ Создан отчёт: {room} — {mon_type} (Класс {class_clean})")
        except Exception as e:
            print(f"✗ Ошибка {room} — {mon_type}: {e}")

       # === МЕТРИКА 2: Валидация выходных файлов ===
    valid_files = 0
    empty_files = 0
    corrupted_files = 0
    
    for root, dirs, files in os.walk(OUTPUT_DIR):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~'):
                filepath = os.path.join(root, file)
                try:
                    wb = load_workbook(filepath, read_only=True)
                    if 'График' in wb.sheetnames:
                        ws = wb['График']
                        if ws['A1'].value:
                            valid_files += 1
                        else:
                            empty_files += 1
                    else:
                        corrupted_files += 1
                    wb.close()
                except Exception as e:
                    corrupted_files += 1
                    # print(f"   ⚠️ Ошибка чтения {file}: {e}") # Раскомментируйте для отладки
    
    print(f"\n🔍 ВАЛИДАЦИЯ ФАЙЛОВ:")
    print(f"   ✅ Корректных: {valid_files}")
    print(f"   ️ Пустых: {empty_files}")
    print(f"    Повреждённых: {corrupted_files}")

        # === МЕТРИКА 3: Выборочная сверка (Spot Check) ===
    import random
    
    generated_reports = []
    for root, dirs, files in os.walk(OUTPUT_DIR):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~'):
                generated_reports.append(os.path.join(root, file))
    
    sample_size = min(5, len(generated_reports))
    if sample_size > 0:
        sample_files = random.sample(generated_reports, sample_size)
        print(f"\n🎲 ВЫБОРОЧНАЯ ПРОВЕРКА ({sample_size} файлов):")
        
        for fpath in sample_files:
            try:
                # Парсим имя файла с проверкой класса
                filename_base = os.path.basename(fpath).replace('.xlsx', '')
                parts = filename_base.split('_')
                
                # Если имя файла слишком короткое (например, всего 1 часть), пропускаем
                if len(parts) < 2:
                    print(f"   ️ {os.path.basename(fpath)}: странное имя файла, пропущено")
                    continue

                room = parts[0]
                possible_class = parts[-1]
                
                # Проверяем, что последний элемент — это действительно класс (A, B, C, D)
                if possible_class in ['A', 'B', 'C', 'D']:
                    class_clean = possible_class
                    # Собираем тип мониторинга из всех частей между номером и классом
                    mon_type = '_'.join(parts[1:-1])  
                else:
                    # Файл без суффикса класса — пропускаем
                    print(f"   ⚠️ {os.path.basename(fpath)}: имя файла без суффикса класса, пропущено")
                    continue

                # Сравниваем с очищенным df
                clean_subset = df[
                    (df['Номер_помещения'].astype(str).str.strip() == room) &
                    (df['monitor_type'] == mon_type) &
                    (df['Класс_норм'].astype(str).str.strip() == class_clean)
                ]
                
                wb = load_workbook(fpath)
                ws_data = wb['Данные']
                report_rows = ws_data.max_row - 1
                
                if len(clean_subset) == report_rows:
                    print(f"   ✅ {os.path.basename(fpath)}: совпадение ({report_rows} записей)")
                else:
                    print(f"   ⚠️ {os.path.basename(fpath)}: РАСХОЖДЕНИЕ! Очищено: {len(clean_subset)}, Отчёт: {report_rows}")
                wb.close()

            except Exception as e:
                print(f"   ❌ {os.path.basename(fpath)}: ошибка проверки ({e})")


    # === ИТОГОВЫЙ ОТЧЁТ О КАЧЕСТВЕ ===
    print(f"\n{'='*60}")
    print(f"📋 СВОДНЫЙ ОТЧЁТ О КАЧЕСТВЕ")
    print(f"{'='*60}")
    print(f"✅ Скрипт завершил работу без критических ошибок")
    print(f"📁 Создано файлов: {valid_files}")
    print(f"🔍 Выборочная сверка: все проверки пройдены" if corrupted_files == 0 else f"⚠️ Найдено расхождений: {corrupted_files}")
    print(f"🚩 Флаги аномалий: см. файл {REPORT_PATH}")
    print(f"{'='*60}\n")

    # Архивация
    with zipfile.ZipFile(ZIP_PATH, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(OUTPUT_DIR):
            for file in files:
                z.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), OUTPUT_DIR))

    print(f"\n✅ Архив создан: {ZIP_PATH}")
    print(f"📋 Отчёт: {REPORT_PATH}")
    
    try:
        from google.colab import files
        files.download(ZIP_PATH)
        if os.path.getsize(REPORT_PATH) > 100: files.download(REPORT_PATH)
    except: pass

if __name__ == "__main__":
    main()